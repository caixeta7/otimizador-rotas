"""
RotaHub - Parser de importação (RF003 + RF004)

Suporta dois formatos reais, identificados por engenharia reversa nas planilhas
fornecidas pelo usuário:

1) "shopee_raw"  — export do app Shopee/Rota Shopee.
   Colunas: AT ID, Sequence, Stop, SPX TN, Destination Address, Bairro, City,
            Zipcode/Postal code, Latitude, Longitude
   - 1 linha = 1 pacote (SPX TN único)
   - Pode ter uma linha com Sequence='-' e Stop='-' = ponto de origem/depósito
   - O campo "Stop" é apenas uma SUGESTÃO de agrupamento da Shopee e é
     inconsistente (comprovado: 25 de 48 grupos tinham coordenadas diferentes
     dentro do mesmo "Stop"). Por isso NÃO usamos esse campo para agrupar -
     agrupamos por proximidade geográfica real (geo.cluster_by_proximity).
   - O complemento/observação vem embutido no próprio texto do endereço,
     separado por vírgula (ex: "Av X, 110, Bl d ap 34").

2) "circuit_processed" — export processado pelo Circuit Route Planner.
   Colunas: AT ID, Destination Address, Bairro, City, Zipcode/Postal code,
            Latitude, Longitude, Address Line 2, Pacotes na Parada
   - 1 linha = 1 parada (já agrupada pelo Circuit)
   - "Address Line 2" é o complemento, já separado
   - "Pacotes na Parada" é uma string irregular com IDs de pacote, ex:
       "117 (+2), 47"
       "6, 7, 8, 9, 11 - Total: 5 pacotes"
     -> extraímos todos os números via regex.
"""
import re
from openpyxl import load_workbook
from .geo import cluster_by_proximity

PACKAGE_NUM_RE = re.compile(r"\d+")
TOTAL_SUFFIX_RE = re.compile(r"-\s*Total:\s*\d+\s*pacotes", re.IGNORECASE)
MAIN_PLUS_EXTRA_RE = re.compile(r"^(\d+)\s*\(\+(\d+)\)$")
PLAIN_NUM_RE = re.compile(r"^(\d+)$")



class ParseError(Exception):
    pass


def _normalize_header(h):
    return (h or "").strip().lower()


def detect_format(headers):
    norm = [_normalize_header(h) for h in headers]
    if "spx tn" in norm and "sequence" in norm:
        return "shopee_raw"
    if "pacotes na parada" in norm:
        return "circuit_processed"
    raise ParseError(
        "Formato de planilha não reconhecido. Colunas encontradas: " + ", ".join(headers)
    )


def _col_index(headers, *candidates):
    norm = [_normalize_header(h) for h in headers]
    for cand in candidates:
        cand_n = _normalize_header(cand)
        if cand_n in norm:
            return norm.index(cand_n)
    return None


def _split_address(raw_address):
    """
    Heurística para separar rua+numero de complemento/observação, já que no
    arquivo bruto da Shopee isso vem tudo junto separado por vírgula.
    Ex: "Rua X, 811" -> ("Rua X, 811", None)
        "Av Nuno M Pereira, 110, Bl d ap 34" -> ("Av Nuno M Pereira, 110", "Bl d ap 34")
    """
    parts = [p.strip() for p in raw_address.split(",")]
    if len(parts) <= 2:
        return raw_address.strip(), None
    street_and_number = ", ".join(parts[:2])
    complement = ", ".join(parts[2:])
    return street_and_number, complement


def _parse_package_ids(raw_value):
    """
    Extrai a lista de pacotes da string 'Pacotes na Parada'.
    Trata 3 padrões observados na planilha real:
      "116 (+1)"                          -> id "116" + 1 pacote extra sem ID
      "6, 7, 8, 9, 11 - Total: 5 pacotes"  -> 5 ids explícitos (sufixo "Total" é só validação)
      "1"                                  -> 1 pacote simples
    IMPORTANTE: "(+N)" é uma CONTAGEM de pacotes extras, não um ID adicional -
    um regex ingênuo de "todos os números" superestima o total (bug encontrado
    e corrigido durante os testes deste parser).
    """
    if raw_value is None:
        return []
    s = str(raw_value)
    expected_total = None
    total_match = re.search(r"Total:\s*(\d+)\s*pacotes", s, re.IGNORECASE)
    if total_match:
        expected_total = int(total_match.group(1))
    s_clean = TOTAL_SUFFIX_RE.sub("", s)

    package_ids = []
    for token in [t.strip() for t in s_clean.split(",") if t.strip()]:
        m = MAIN_PLUS_EXTRA_RE.match(token)
        if m:
            main_id, extra = m.group(1), int(m.group(2))
            package_ids.append(main_id)
            package_ids.extend(f"{main_id}+{i+1}" for i in range(extra))
            continue
        m2 = PLAIN_NUM_RE.match(token)
        if m2:
            package_ids.append(m2.group(1))

    if expected_total is not None and len(package_ids) != expected_total:
        diff = expected_total - len(package_ids)
        if diff > 0:
            package_ids.extend(f"extra{i+1}" for i in range(diff))
        elif diff < 0:
            package_ids = package_ids[:expected_total]

    return package_ids


def parse_workbook(file_path):
    """
    Lê o arquivo .xlsx e retorna um dict normalizado:
    {
        "format": "shopee_raw" | "circuit_processed",
        "at_id": str | None,
        "origin": {"latitude":..., "longitude":..., "address":...} | None,
        "stops": [
            {
                "address": str, "complement": str|None, "neighborhood": str,
                "city": str, "zipcode": str, "latitude": float, "longitude": float,
                "needs_review": bool, "packages": [ "tracking_or_id", ... ]
            }, ...
        ]
    }
    """
    wb = load_workbook(file_path, data_only=True)
    ws = wb[wb.sheetnames[0]]

    rows_iter = ws.iter_rows(values_only=True)
    try:
        headers = list(next(rows_iter))
    except StopIteration:
        raise ParseError("Planilha vazia.")

    fmt = detect_format(headers)

    raw_rows = []
    for row in rows_iter:
        if row is None or all(c is None for c in row):
            continue
        raw_rows.append(row)

    at_id = None
    origin = None
    prelim_points = []  # pontos "crus" antes do agrupamento (1 por linha/pacote)

    if fmt == "shopee_raw":
        idx = {
            "at_id": _col_index(headers, "AT ID"),
            "sequence": _col_index(headers, "Sequence"),
            "stop": _col_index(headers, "Stop"),
            "tn": _col_index(headers, "SPX TN"),
            "address": _col_index(headers, "Destination Address"),
            "bairro": _col_index(headers, "Bairro"),
            "city": _col_index(headers, "City"),
            "zip": _col_index(headers, "Zipcode/Postal code", "Zipcode"),
            "lat": _col_index(headers, "Latitude"),
            "lng": _col_index(headers, "Longitude"),
        }
        required = ["address", "lat", "lng"]
        for r in required:
            if idx[r] is None:
                raise ParseError(f"Coluna obrigatória ausente no formato Shopee bruto: {r}")

        # NOTA: linhas com Sequence='-' e Stop='-' foram inicialmente tratadas
        # como "ponto de origem/depósito" (baseado num único exemplo de
        # planilha). Isso se mostrou ERRADO: em planilhas com múltiplas
        # linhas assim, elas são entregas normais (endereços residenciais
        # reais, com rastreio próprio), só ainda sem posição definida pela
        # Shopee - confirmado comparando com o arquivo já processado pelo
        # Circuit, que trata essas mesmas linhas como paradas comuns. Por
        # isso agora tratamos '-' apenas como "sequência ainda não definida"
        # e incluímos a linha normalmente na lista de entregas.
        for row in raw_rows:
            at_id = at_id or row[idx["at_id"]]
            raw_address = row[idx["address"]]
            lat = row[idx["lat"]]
            lng = row[idx["lng"]]

            if raw_address is None:
                continue

            street, complement = _split_address(str(raw_address))
            tracking = row[idx["tn"]] if idx["tn"] is not None else None

            prelim_points.append({
                "address": street,
                "complement": complement,
                "neighborhood": row[idx["bairro"]] if idx["bairro"] is not None else None,
                "city": row[idx["city"]] if idx["city"] is not None else None,
                "zipcode": row[idx["zip"]] if idx["zip"] is not None else None,
                "latitude": float(lat) if lat is not None else None,
                "longitude": float(lng) if lng is not None else None,
                "needs_review": lat is None or lng is None,
                "package_ref": str(tracking) if tracking else None,
            })

    elif fmt == "circuit_processed":
        idx = {
            "at_id": _col_index(headers, "AT ID"),
            "address": _col_index(headers, "Destination Address"),
            "bairro": _col_index(headers, "Bairro"),
            "city": _col_index(headers, "City"),
            "zip": _col_index(headers, "Zipcode/Postal code", "Zipcode"),
            "lat": _col_index(headers, "Latitude"),
            "lng": _col_index(headers, "Longitude"),
            "line2": _col_index(headers, "Address Line 2"),
            "packages": _col_index(headers, "Pacotes na Parada"),
        }
        required = ["address", "lat", "lng"]
        for r in required:
            if idx[r] is None:
                raise ParseError(f"Coluna obrigatória ausente no formato Circuit: {r}")

        for row in raw_rows:
            at_id = at_id or row[idx["at_id"]]
            raw_address = row[idx["address"]]
            lat = row[idx["lat"]]
            lng = row[idx["lng"]]
            if raw_address is None:
                continue

            pkg_ids = _parse_package_ids(row[idx["packages"]] if idx["packages"] is not None else None)
            if not pkg_ids:
                pkg_ids = [None]  # ao menos 1 pacote "genérico" nessa parada

            for pkg_id in pkg_ids:
                prelim_points.append({
                    "address": str(raw_address).strip(),
                    "complement": row[idx["line2"]] if idx["line2"] is not None else None,
                    "neighborhood": row[idx["bairro"]] if idx["bairro"] is not None else None,
                    "city": row[idx["city"]] if idx["city"] is not None else None,
                    "zipcode": row[idx["zip"]] if idx["zip"] is not None else None,
                    "latitude": float(lat) if lat is not None else None,
                    "longitude": float(lng) if lng is not None else None,
                    "needs_review": lat is None or lng is None,
                    "package_ref": pkg_id,
                })

    if not prelim_points:
        raise ParseError("Nenhuma parada válida encontrada na planilha.")

    # Pontos sem coordenada (RF005) não podem entrar no clustering geográfico
    # - viram parada própria, marcada para revisão/geocodificação manual.
    geocoded_points = [p for p in prelim_points if p["latitude"] is not None]
    missing_points = [p for p in prelim_points if p["latitude"] is None]

    # Agrupa por proximidade geográfica real (RF004/RF007) - não pelo campo "Stop"
    cluster_ids = cluster_by_proximity(geocoded_points) if geocoded_points else []

    clusters = {}
    for point, cid in zip(geocoded_points, cluster_ids):
        clusters.setdefault(cid, []).append(point)
    next_cid = (max(clusters.keys()) + 1) if clusters else 0
    for missing_point in missing_points:
        clusters[next_cid] = [missing_point]
        next_cid += 1

    stops = []
    for cid, pts in clusters.items():
        first = pts[0]
        has_coords = pts[0]["latitude"] is not None
        packages = [p["package_ref"] for p in pts if p.get("package_ref")]
        # se paradas do mesmo cluster têm endereços de texto diferentes
        # (prédios vizinhos), guardamos o mais frequente como principal.
        addresses = [p["address"] for p in pts]
        main_address = max(set(addresses), key=addresses.count)
        # IMPORTANTE: a coordenada da parada é a coordenada REAL de um ponto
        # que tem esse endereço principal - não a média de todo o cluster.
        # A média pode cair num ponto sintético que não corresponde a
        # endereço nenhum (ex: no meio da rua, ou puxada pro prédio vizinho),
        # o que manda a navegação (Google Maps) pro lugar errado. Bug real
        # relatado pelo usuário e corrigido aqui.
        if has_coords:
            representative = next(p for p in pts if p["address"] == main_address)
            rep_lat, rep_lng = representative["latitude"], representative["longitude"]
        else:
            rep_lat, rep_lng = None, None
        # demais campos (complemento, bairro, cep) tambem vem do ponto
        # representativo - nao do "primeiro" ponto arbitrario do cluster,
        # pelo mesmo motivo: podem pertencer a outro endereco do grupo.
        rep_for_fields = next((p for p in pts if p["address"] == main_address), first)
        stops.append({
            "address": main_address,
            "complement": rep_for_fields.get("complement"),
            "neighborhood": rep_for_fields.get("neighborhood"),
            "city": rep_for_fields.get("city"),
            "zipcode": rep_for_fields.get("zipcode"),
            "latitude": rep_lat,
            "longitude": rep_lng,
            "needs_review": not has_coords,
            "packages": packages if packages else [f"PKG-{cid+1}"],
        })

    return {
        "format": fmt,
        "at_id": str(at_id) if at_id else None,
        "origin": origin,
        "stops": stops,
    }
