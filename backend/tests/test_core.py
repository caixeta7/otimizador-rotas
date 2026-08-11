"""
Testes automatizados do RotaHub - roda com: pytest backend/tests/

Cobre as validações feitas manualmente durante o desenvolvimento:
- Parser não perde nem duplica dados nos 2 formatos reais suportados.
- RF005: endereço sem coordenada é sinalizado, nunca descartado.
- RF006: o otimizador realmente evita o padrão de "volta desnecessária"
  relatado pelo usuário, e é determinístico o suficiente para não regredir.
"""
import os
import sys
import math
import pytest

sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))

from app.parser import parse_workbook, _parse_package_ids, ParseError
from app.optimizer import solve_tsp, route_total_distance
from app.distance import _haversine_matrix
from app.geo import cluster_by_proximity, haversine

FIXTURES = os.path.join(os.path.dirname(__file__), "fixtures")


# ------------------------------------------------------------- parser -----

def test_detects_circuit_format():
    result = parse_workbook(os.path.join(FIXTURES, "exemplo_circuit_processado.xlsx"))
    assert result["format"] == "circuit_processed"
    assert result["at_id"] == "AT202607096ZRCP"


def test_detects_shopee_raw_format():
    result = parse_workbook(os.path.join(FIXTURES, "exemplo_shopee_bruto.xlsx"))
    assert result["format"] == "shopee_raw"


def test_circuit_no_data_loss():
    """46 paradas / 129 pacotes é o número validado manualmente linha a linha."""
    result = parse_workbook(os.path.join(FIXTURES, "exemplo_circuit_processado.xlsx"))
    total_packages = sum(len(s["packages"]) for s in result["stops"])
    assert len(result["stops"]) == 46
    assert total_packages == 129


def test_shopee_raw_no_data_loss():
    """
    113 linhas na planilha = 113 pacotes reais. As linhas com Sequence='-'
    NÃO são um depósito (suposição inicial errada, corrigida depois de
    comparar com a planilha da Bruna) - são entregas normais ainda sem
    sequência definida pela Shopee, e por isso entram na contagem.
    """
    result = parse_workbook(os.path.join(FIXTURES, "exemplo_shopee_bruto.xlsx"))
    total_packages = sum(len(s["packages"]) for s in result["stops"])
    assert total_packages == 113


def test_bruna_files_no_data_loss():
    """
    Planilha da Bruna tem DUAS linhas com Sequence='-' (não uma, como a do
    Paulo) - confirma que não é um caso especial de depósito único, e sim o
    padrão real: pacotes ainda sem sequência atribuída.
    """
    circuit = parse_workbook(os.path.join(FIXTURES, "exemplo_bruna_circuit_processado.xlsx"))
    assert sum(len(s["packages"]) for s in circuit["stops"]) == 23

    raw = parse_workbook(os.path.join(FIXTURES, "exemplo_bruna_shopee_bruto.xlsx"))
    assert sum(len(s["packages"]) for s in raw["stops"]) == 20
    assert raw["origin"] is None


def test_package_field_parsing_edge_cases():
    assert _parse_package_ids("116 (+1)") == ["116", "116+1"]
    assert _parse_package_ids("1") == ["1"]
    ids = _parse_package_ids("6, 7, 8, 9, 11 - Total: 5 pacotes")
    assert len(ids) == 5
    assert _parse_package_ids(None) == []


def test_missing_coordinates_are_flagged_not_dropped(tmp_path):
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Planilha Processada"
    ws.append(["AT ID", "Destination Address", "Bairro", "City", "Zipcode/Postal code",
               "Latitude", "Longitude", "Address Line 2", "Pacotes na Parada"])
    ws.append(["AT1", "Rua Sem Coordenada, 1", "Centro", "Sao Paulo", "01000-000",
               None, None, None, "1"])
    path = tmp_path / "sem_geo.xlsx"
    wb.save(path)

    result = parse_workbook(str(path))
    assert len(result["stops"]) == 1
    assert result["stops"][0]["needs_review"] is True


def test_stop_coordinate_is_real_point_not_average(tmp_path):
    """
    Regressão: a coordenada de uma parada agrupada precisa ser a coordenada
    REAL de um dos pontos originais (o do endereço principal), nunca a média
    do cluster. Uma média pode cair num ponto sintético que não corresponde
    a endereço nenhum - bug real que mandava a navegação (Google Maps) pra
    rua errada.
    """
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(["AT ID", "Sequence", "Stop", "SPX TN", "Destination Address", "Bairro",
               "City", "Zipcode/Postal code", "Latitude", "Longitude"])
    # dois enderecos DIFERENTES mas proximos o bastante pra cair no mesmo
    # cluster (tolerancia de 25m) - o principal (2 ocorrencias) deve "vencer"
    ws.append(["AT1", 1, 1, "TN1", "Rua A, 100", "Centro", "Sao Paulo", "01000-000",
               -23.55000, -46.63000])
    ws.append(["AT1", 2, 1, "TN2", "Rua A, 100", "Centro", "Sao Paulo", "01000-000",
               -23.55002, -46.63002])
    ws.append(["AT1", 3, 1, "TN3", "Rua B, 50", "Centro", "Sao Paulo", "01000-000",
               -23.54999, -46.62999])
    path = tmp_path / "cluster_test.xlsx"
    wb.save(path)

    result = parse_workbook(str(path))
    assert len(result["stops"]) == 1
    stop = result["stops"][0]
    assert stop["address"] == "Rua A, 100"  # maioria (2 de 3 pontos)
    # a coordenada tem que ser EXATAMENTE uma das duas coordenadas reais de
    # "Rua A, 100" - nunca a media dos 3 pontos do cluster.
    real_coords_rua_a = {(-23.55000, -46.63000), (-23.55002, -46.63002)}
    assert (round(stop["latitude"], 5), round(stop["longitude"], 5)) in real_coords_rua_a


def test_empty_spreadsheet_raises():
    import openpyxl
    import tempfile
    import os
    wb = openpyxl.Workbook()
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
        path = f.name
    try:
        wb.save(path)
        with pytest.raises(ParseError):
            parse_workbook(path)
    finally:
        try:
            os.remove(path)
        except (OSError, PermissionError):
            pass


# --------------------------------------------------------------- geo ------

def test_cluster_by_proximity_groups_close_points():
    points = [
        {"latitude": -23.6500, "longitude": -46.7400},
        {"latitude": -23.65001, "longitude": -46.74001},  # ~1m de distancia
        {"latitude": -23.7000, "longitude": -46.8000},     # bem longe
    ]
    clusters = cluster_by_proximity(points, tolerance_m=25)
    assert clusters[0] == clusters[1]
    assert clusters[2] != clusters[0]


def test_cluster_by_proximity_no_chaining_effect():
    """
    Regressão: um centróide que se desloca a cada ponto novo permite um
    "efeito cadeia" (A-B a 20m, centro desloca, C a 20m do NOVO centro mas
    a 35-40m do A original também entra no grupo) - isso pode juntar
    endereços de ruas diferentes numa mesma parada mesmo respeitando 25m a
    cada passo. Com âncora fixa (o primeiro ponto do grupo), todo ponto tem
    que estar a no máximo `tolerance_m` do ponto que abriu o grupo - sem
    exceção. Este teste garante essa invariante com 10 pontos em fila,
    15m um do outro (100+m de ponta a ponta).
    """
    lat0, lng0 = -23.55000, -46.63000
    step_m = 15
    points = []
    for i in range(10):
        dlng = (step_m * i) / 111320 / math.cos(math.radians(lat0))
        points.append({"latitude": lat0, "longitude": lng0 + dlng})

    clusters = cluster_by_proximity(points, tolerance_m=25)

    anchors = {}
    for i, c in enumerate(clusters):
        if c not in anchors:
            anchors[c] = points[i]
        d_m = haversine(points[i]["latitude"], points[i]["longitude"],
                         anchors[c]["latitude"], anchors[c]["longitude"]) * 1000
        assert d_m <= 25.01, f"ponto {i} está a {d_m:.1f}m da âncora do cluster {c} (efeito cadeia)"


# ---------------------------------------------------------- optimizer -----

def test_tsp_avoids_unnecessary_backtrack():
    """
    Reproduz o problema relatado pelo usuário: 2 paradas na mesma rua
    (extremos da lista) e outras paradas espalhadas em outra direção.
    A rota otimizada deve visitar as 2 paradas da mesma rua em sequência.
    """
    points = [
        {"latitude": -23.6500, "longitude": -46.7400},  # 0 origem
        {"latitude": -23.6501, "longitude": -46.7401},  # 1 Rua X - parada 1
        {"latitude": -23.6600, "longitude": -46.7500},  # 2 bairro B
        {"latitude": -23.6650, "longitude": -46.7550},  # 3 bairro B
        {"latitude": -23.6700, "longitude": -46.7600},  # 4 bairro B
        {"latitude": -23.6502, "longitude": -46.7402},  # 5 Rua X - parada 50
    ]
    matrix = _haversine_matrix(points)
    order = solve_tsp(matrix, start_index=0, time_limit_seconds=5)

    assert sorted(order) == list(range(len(points)))  # visita todo mundo 1x
    pos_1 = order.index(1)
    pos_5 = order.index(5)
    assert abs(pos_1 - pos_5) == 1, "as duas paradas da mesma rua deveriam ficar adjacentes na rota"


def test_tsp_beats_naive_order_on_real_data():
    """A rota otimizada nunca deve ser pior que a ordem crua de importação."""
    result = parse_workbook(os.path.join(FIXTURES, "exemplo_shopee_bruto.xlsx"))
    points = result["stops"]
    matrix = _haversine_matrix(points)

    naive_order = list(range(len(points)))
    naive_dist = route_total_distance(naive_order, matrix)

    opt_order = solve_tsp(matrix, start_index=0, time_limit_seconds=10)
    opt_dist = route_total_distance(opt_order, matrix)

    assert opt_dist <= naive_dist
    assert opt_dist < naive_dist * 0.9  # pelo menos 10% melhor nesse dataset


def test_tsp_single_and_two_points_dont_crash():
    assert solve_tsp([[0]], start_index=0) == [0]
    assert solve_tsp([[0, 1], [1, 0]], start_index=0) == [0, 1]
